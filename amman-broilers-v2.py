import os
import time
import tkinter as tk
from tkinter import messagebox
from fpdf import FPDF
import openpyxl

class BillingSystem:
    def __init__(self, root):
        self.root = root
        self.root.title("Amman Broilers")
        self.root.geometry("600x600")
        self.root.config(bg="#f4f4f4")  

        self.header_frame = tk.Frame(self.root, bg="#4CAF50", pady=20)
        self.header_frame.pack(fill=tk.X)

        self.header_label = tk.Label(self.header_frame, text="Amman Broilers", font=("Helvetica", 24, 'bold'), fg="white", bg="#4CAF50")
        self.header_label.pack()

        self.items = {
            'Live Broiler': 150,
            'Broiler Chicken': 220,
            'Chicken Wings': 240,
            'Chicken Leg': 240,
            'Chicken Lollipop': 240,
            'Boneless': 400,
            'Live Country': 680,
            'Country Chicken': 850
        }

        self.current_total = 0
        self.serial_number = 1
        self.bill_items = []

        self.excel_path = r"D:\Bill\sales_data.xlsx"
        self.sheet = None
        self.workbook = None

        if os.path.exists(self.excel_path):
            self.workbook = openpyxl.load_workbook(self.excel_path)
            self.sheet = self.workbook.active
        else:
            self.workbook = openpyxl.Workbook()
            self.sheet = self.workbook.active
            self.sheet.append(["Serial No.", "Customer Name", "Item", "Weight (kg)", "Total Amount", "Timestamp"])
            self.workbook.save(self.excel_path)

        self.serial_number += 1

        self.customer_name_label = tk.Label(self.root, text="Customer Name:", font=("Arial", 12))
        self.customer_name_label.pack(pady=10)
        self.customer_name = tk.Entry(self.root, font=("Arial", 12), width=25)
        self.customer_name.pack(pady=5)

        self.items_combobox = tk.StringVar(self.root)
        self.items_combobox.set('Select Item')
        self.items_menu = tk.OptionMenu(self.root, self.items_combobox, *self.items.keys())
        self.items_menu.config(font=("Arial", 12), width=25)
        self.items_menu.pack(pady=10)

        self.weight_label = tk.Label(self.root, text="Weight (kg):", font=("Arial", 12))
        self.weight_label.pack(pady=10)
        self.weight = tk.Entry(self.root, font=("Arial", 12), width=10)
        self.weight.pack(pady=5)

        self.total_label = tk.Label(self.root, text="Total: INR 0", font=("Arial", 14, 'bold'))
        self.total_label.pack(pady=20)

        self.add_button = tk.Button(self.root, text="Add to Bill", font=("Arial", 14, 'bold'), bg="#4CAF50", fg="white", relief="flat", command=self.add_to_bill)
        self.add_button.pack(pady=10, fill=tk.X, padx=50)

        self.generate_button = tk.Button(self.root, text="Generate Receipt", font=("Arial", 14, 'bold'), bg="#007BFF", fg="white", relief="flat", command=self.generate_receipt)
        self.generate_button.pack(pady=10, fill=tk.X, padx=50)

        self.status_label = tk.Label(self.root, text="", font=("Arial", 12), fg="green", bg="#f4f4f4")
        self.status_label.pack(pady=10)

    def get_serial_number(self):
        last_row = len(self.sheet['A'])
        return last_row + 1

    def add_to_bill(self):
        item = self.items_combobox.get()
        weight = self.weight.get()

        if item == 'Select Item' or not self.is_valid_weight(weight):
            messagebox.showerror("Error", "Please select a valid item and weight (in kg).")
            return

        weight = float(weight)
        item_price = self.items.get(item)
        total_price = item_price * weight

        self.bill_items.append((item, weight, item_price, total_price))
        self.current_total += total_price
        self.total_label.config(text=f"Total: INR {self.current_total:.2f}")

        self.items_combobox.set('Select Item')
        self.weight.delete(0, tk.END)

    def is_valid_weight(self, weight):
        try:
            weight_value = float(weight)
            return weight_value > 0
        except ValueError:
            return False

    def generate_receipt(self):
        customer_name = self.customer_name.get()
        if not customer_name:
            messagebox.showerror("Error", "Customer Name is required.")
            return

        pdf = FPDF()
        pdf.add_page()

        pdf.set_font("Arial", size=16, style='B')
        pdf.cell(200, 10, txt="Amman Broilers", ln=True, align='C')
        pdf.set_font("Arial", size=12, style='B')
        pdf.cell(200, 10, txt="***** BILL *****", ln=True, align='C')
        pdf.ln(10)

        pdf.set_font("Arial", size=12, style='B')
        pdf.cell(50, 10, txt="Customer Name:", border=1, align='L')
        pdf.cell(140, 10, txt=customer_name, border=1)
        pdf.ln(10)

        pdf.set_font("Arial", size=12, style='B')
        pdf.cell(50, 10, txt="Item", border=1, align='C')
        pdf.cell(50, 10, txt="Weight (kg)", border=1, align='C')
        pdf.cell(40, 10, txt="Price (INR)", border=1, align='C')
        pdf.cell(50, 10, txt="Total (INR)", border=1, align='C')
        pdf.ln(10)

        for item, weight, item_price, total_price in self.bill_items:
            pdf.cell(50, 10, txt=item, border=1, align='C')
            pdf.cell(50, 10, txt=str(weight), border=1, align='C')
            pdf.cell(40, 10, txt=str(item_price), border=1, align='C')
            pdf.cell(50, 10, txt=str(total_price), border=1, align='C')
            pdf.ln(10)

        pdf.ln(10)
        pdf.set_font("Arial", size=12, style='B')
        pdf.cell(50, 10, txt="Total (INR):", border=1, align='C')
        pdf.cell(140, 10, txt=f"{self.current_total:.2f}", border=1)
        pdf.ln(10)

        timestamp_date = time.strftime("%d-%m-%Y")
        timestamp_time = time.strftime("%I:%M:%S %p")

        pdf.cell(50, 10, txt="Date:", border=1, align='C')
        pdf.cell(140, 10, txt=timestamp_date, border=1)
        pdf.ln(10)

        pdf.cell(50, 10, txt="Time:", border=1, align='C')
        pdf.cell(140, 10, txt=timestamp_time, border=1)
        pdf.ln(10)

        pdf.rect(10, 10, 190, 260)

        receipt_filename = f"{self.serial_number}-{customer_name}.pdf"
        receipt_path = os.path.join(r"D:\Bill", receipt_filename)

        directory = r"D:\Bill"
        if not os.path.exists(directory):
            try:
                os.makedirs(directory)
            except PermissionError:
                messagebox.showerror("Error", f"Permission denied: Cannot create directory '{directory}'.")
                return

        try:
            pdf.output(receipt_path)
            messagebox.showinfo("Receipt", f"Receipt generated successfully! Saved as {receipt_filename}")
        except PermissionError:
            messagebox.showerror("Error", f"Permission denied: Unable to save the file to {receipt_path}.")
            return

        for item, weight, item_price, total_price in self.bill_items:
            self.sheet.append([self.serial_number, customer_name, item, weight, total_price, timestamp_time])

        self.workbook.save(self.excel_path)

        self.serial_number += 1
        self.current_total = 0
        self.bill_items.clear()
        self.total_label.config(text="Total: INR 0")
        self.customer_name.delete(0, tk.END)
        self.status_label.config(text="Bill has been added successfully!", fg="green")

if __name__ == "__main__":
    root = tk.Tk()
    app = BillingSystem(root)
    root.mainloop()
